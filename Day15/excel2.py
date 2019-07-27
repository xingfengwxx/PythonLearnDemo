#!/usr/bin/env python 3.7
# -*- coding: utf-8 -*-
# @Time       : 2019/7/27 10:02
# @Author     : wangxingxing
# @Email      : xingfengwxx@gmail.com 
# @File       : excel2.py
# @Software   : PyCharm
# @Description: 读取Excel文件

from openpyxl import load_workbook

workbook = load_workbook('./res/学生明细表.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
